"""CSV/TSV 摘要与 xlsx 只读抽样（纯本地文件）。"""

from __future__ import annotations

import csv
import os
from typing import Any

from _meta import skill

_MAX_FILE_BYTES = 40 * 1024 * 1024


def _abs_file(path: str) -> str:
    p = os.path.abspath(os.path.expanduser(path))
    if not os.path.isfile(p):
        raise FileNotFoundError(f"文件不存在: {p}")
    if os.path.getsize(p) > _MAX_FILE_BYTES:
        raise ValueError(f"文件过大（>{_MAX_FILE_BYTES // (1024 * 1024)}MB），请缩小或拆分后再试。")
    return p


@skill(
    desc="读取 CSV/TSV 并返回列名、行数、抽样行与数值列简单统计（min/max/mean）。",
    examples=[
        "csv_tsv_summary('./data/metrics.csv')",
        "csv_tsv_summary('./t.tsv', delimiter='\\t', max_rows=2000)",
    ],
)
def csv_tsv_summary(
    path: str,
    delimiter: str = "",
    max_rows: int = 8000,
    sample_rows: int = 5,
) -> dict[str, Any]:
    """delimiter 为空则按扩展名与 Sniffer 猜测分隔符。"""
    abs_path = _abs_file(path)
    max_rows = max(10, min(int(max_rows), 200_000))
    sample_rows = max(0, min(int(sample_rows), 50))
    lower = abs_path.lower()
    delim = (delimiter or "").strip()
    if not delim:
        if lower.endswith(".tsv"):
            delim = "\t"
        else:
            delim = ""

    rows_data: list[list[str]] = []
    with open(abs_path, "r", encoding="utf-8", errors="replace", newline="") as fh:
        sample_text = fh.read(65536)
        fh.seek(0)
        if not delim:
            try:
                dialect = csv.Sniffer().sniff(sample_text, delimiters=",\t;|")
                delim = dialect.delimiter
            except csv.Error:
                delim = ","
        reader = csv.reader(fh, delimiter=delim)
        for _i, row in enumerate(reader):
            if len(rows_data) >= max_rows:
                break
            rows_data.append([str(c) for c in row])

    if not rows_data:
        return {
            "path": abs_path,
            "delimiter_repr": repr(delim),
            "column_names": [],
            "rows_read": 0,
            "data_rows": 0,
            "truncated": False,
            "sample_rows": [],
            "column_stats": [],
        }

    header = rows_data[0]
    body = rows_data[1:]
    rows_read = len(rows_data)
    truncated = rows_read >= max_rows
    sample = body[:sample_rows]
    col_count = len(header)
    stats: list[dict[str, Any]] = []
    stat_rows = body[: min(2000, len(body))]
    for ci in range(col_count):
        vals: list[float] = []
        for r in stat_rows:
            if ci >= len(r):
                continue
            s = (r[ci] or "").strip()
            if not s:
                continue
            try:
                vals.append(float(s.replace(",", "")))
            except ValueError:
                continue
        name = header[ci] if ci < len(header) else f"col_{ci}"
        if vals:
            stats.append(
                {
                    "column": name,
                    "numeric_count": len(vals),
                    "min": min(vals),
                    "max": max(vals),
                    "mean": round(sum(vals) / len(vals), 6),
                }
            )
        else:
            stats.append({"column": name, "numeric_count": 0})

    return {
        "path": abs_path,
        "delimiter_repr": repr(delim),
        "column_names": header,
        "rows_read": rows_read,
        "data_rows": max(0, rows_read - 1),
        "truncated": truncated,
        "sample_rows": sample,
        "column_stats": stats,
    }


@skill(
    desc="只读抽样 xlsx：表名、指定 sheet 的前若干行与列（openpyxl read_only）。",
    examples=[
        "xlsx_sample('./report.xlsx')",
        "xlsx_sample('./book.xlsx', sheet='Sheet1', max_rows=20, max_cols=12)",
    ],
)
def xlsx_sample(
    path: str,
    sheet: str = "0",
    max_rows: int = 40,
    max_cols: int = 24,
) -> dict[str, Any]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("缺少依赖 openpyxl，请安装：pip install openpyxl") from exc

    abs_path = _abs_file(path)
    if not abs_path.lower().endswith((".xlsx", ".xlsm")):
        raise ValueError("仅支持 .xlsx / .xlsm")

    max_rows = max(1, min(int(max_rows), 500))
    max_cols = max(1, min(int(max_cols), 256))

    wb = openpyxl.load_workbook(abs_path, read_only=True, data_only=True)
    try:
        names = list(wb.sheetnames)
        sh_raw = (sheet or "0").strip()
        if sh_raw.isdigit():
            idx = int(sh_raw)
            if idx < 0 or idx >= len(names):
                raise ValueError(f"sheet 索引越界: {idx}，有效 0..{len(names) - 1}")
            ws = wb[names[idx]]
            sheet_used = names[idx]
        else:
            if sh_raw not in names:
                raise ValueError(f"未找到工作表: {sh_raw!r}，已有: {names[:20]}")
            ws = wb[sh_raw]
            sheet_used = sh_raw

        grid: list[list[Any]] = []
        for i, row in enumerate(ws.iter_rows(max_row=max_rows, max_col=max_cols, values_only=True)):
            if i >= max_rows:
                break
            grid.append([("" if c is None else c) for c in row])
    finally:
        wb.close()

    return {
        "path": abs_path,
        "sheet_names": names,
        "sheet_used": sheet_used,
        "max_rows": max_rows,
        "max_cols": max_cols,
        "rows_returned": len(grid),
        "grid": grid,
    }
